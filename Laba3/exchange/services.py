import json
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference


DEFAULT_SCHEMA = [
    {"code": "product_name", "order": 1, "name": "Наименование товара", "type": "string", "precision": 0, "transmit": True},
    {"code": "product_group", "order": 2, "name": "Группа товара", "type": "string", "precision": 0, "transmit": True},
    {"code": "sold_qty", "order": 3, "name": "Проданное количество", "type": "number", "precision": 0, "transmit": True},
    {"code": "sale_price", "order": 4, "name": "Продажная цена", "type": "number", "precision": 2, "transmit": True},
    {"code": "purchase_price", "order": 5, "name": "Закупочная цена", "type": "number", "precision": 2, "transmit": True},
    {"code": "discount", "order": 6, "name": "Скидка (%)", "type": "number", "precision": 2, "transmit": True},
    {"code": "sale_date", "order": 7, "name": "Дата продажи", "type": "date", "precision": 0, "transmit": True},
]


def _save_json(path: str | Path, payload: dict[str, Any]) -> Path:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return p


def _load_json(path: str | Path) -> dict[str, Any]:
    with Path(path).open("r", encoding="utf-8") as f:
        return json.load(f)


def validate_schema(schema: list[dict[str, Any]]) -> None:
    required = {"code", "order", "name", "type", "precision", "transmit"}
    allowed_types = {"string", "date", "number"}

    codes = set()
    orders = set()
    for field in schema:
        if not required.issubset(field.keys()):
            missing = required - set(field.keys())
            raise ValueError(f"Поле схемы не содержит обязательные ключи: {missing}")

        if field["type"] not in allowed_types:
            raise ValueError(f"Неверный тип поля: {field['type']}")

        if field["code"] in codes:
            raise ValueError(f"Дублирующийся code: {field['code']}")
        if field["order"] in orders:
            raise ValueError(f"Дублирующийся order: {field['order']}")
        codes.add(field["code"])
        orders.add(field["order"])


def _cast_value(value: Any, field_type: str, precision: int) -> Any:
    if field_type == "string":
        return "" if value is None else str(value)

    if field_type == "date":
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, str):
            # ожидается YYYY-MM-DD
            return datetime.strptime(value, "%Y-%m-%d").date().isoformat()
        raise ValueError(f"Некорректное значение даты: {value}")

    if field_type == "number":
        d = Decimal(str(value if value is not None else 0))
        quant = Decimal("1") if precision <= 0 else Decimal("1." + "0" * precision)
        return float(d.quantize(quant, rounding=ROUND_HALF_UP))

    raise ValueError(f"Неизвестный тип: {field_type}")


def export_source_payload(
    schema: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    out_path: str | Path
) -> Path:
    validate_schema(schema)
    sorted_schema = sorted(schema, key=lambda x: x["order"])

    normalized_rows: list[dict[str, Any]] = []
    for row in rows:
        out_row: dict[str, Any] = {}
        for field in sorted_schema:
            if not field.get("transmit", True):
                continue
            code = field["code"]
            out_row[code] = _cast_value(row.get(code), field["type"], int(field.get("precision", 0)))
        normalized_rows.append(out_row)

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "schema": sorted_schema,
        "rows": normalized_rows,
    }
    return _save_json(out_path, payload)


def process_profit_by_group(
    source_payload_path: str | Path,
    out_path: str | Path
) -> Path:
    payload = _load_json(source_payload_path)
    rows = payload.get("rows", [])

    required_codes = {"product_group", "sold_qty", "sale_price", "purchase_price"}
    if rows:
        missing_codes = required_codes - set(rows[0].keys())
        if missing_codes:
            raise ValueError(f"В данных отсутствуют обязательные поля: {missing_codes}")

    totals: dict[str, Decimal] = {}

    for row in rows:
        group = str(row.get("product_group", "Без группы"))
        qty = Decimal(str(row.get("sold_qty", 0)))
        sale_price = Decimal(str(row.get("sale_price", 0)))
        purchase_price = Decimal(str(row.get("purchase_price", 0)))

        # По условию: прибыль = цена продажи (без скидки) * количество - закупочная * количество
        profit = (sale_price * qty) - (purchase_price * qty)
        totals[group] = totals.get(group, Decimal("0")) + profit

    result = {
        "source_file": str(source_payload_path),
        "processed_at": datetime.now().isoformat(timespec="seconds"),
        "groups": [
            {"product_group": grp, "profit": float(val.quantize(Decimal("1.00")))}
            for grp, val in sorted(totals.items(), key=lambda x: x[0])
        ],
    }
    return _save_json(out_path, result)


def build_excel_visualization(
    processed_payload_path: str | Path,
    out_xlsx_path: str | Path
) -> Path:
    data = _load_json(processed_payload_path)
    groups = data.get("groups", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Прибыль по группам"

    ws["A1"] = "Группа товара"
    ws["B1"] = "Сумма прибыли"

    row_idx = 2
    for item in groups:
        ws.cell(row=row_idx, column=1, value=item["product_group"])
        ws.cell(row=row_idx, column=2, value=item["profit"])
        row_idx += 1

    if row_idx > 2:
        chart = BarChart()
        chart.title = "Динамика прибыли по группам"
        chart.y_axis.title = "Прибыль"
        chart.x_axis.title = "Группа"

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=row_idx - 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=row_idx - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 8
        chart.width = 16

        ws.add_chart(chart, "D2")

    out_path = Path(out_xlsx_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path